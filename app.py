from flask import Flask, render_template, request, jsonify, send_file
from engine.predictor import predict_attendance
from engine.economics import calculate_waste_cost
from engine.gate import transparency_gate
from engine.student import get_student_insights
from engine.mailer import send_reminder_email
import csv
import io
import os

# Load .env file if it exists
def _load_env():
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip())

_load_env()

app = Flask(__name__)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/organizer")
def organizer():
    return render_template("organizer.html")

@app.route("/student")
def student():
    return render_template("student.html")

@app.route("/api/predict", methods=["POST"])
def predict():
    data = request.get_json()
    gate_result = transparency_gate(data)
    if gate_result["status"] == "BLOCKED":
        return jsonify({"error": gate_result["reason"]}), 400

    prediction = predict_attendance(data)
    economics = calculate_waste_cost(
        planned=data.get("planned_quantity", 0),
        predicted=prediction["predicted_attendance"],
        event_type=data.get("event_type", "general"),
        cost_per_person=data.get("cost_per_person", 15)
    )
    return jsonify({**prediction, **economics})


@app.route("/api/upload", methods=["POST"])
def upload_file():
    """
    Accept CSV or Excel file upload.
    Returns predictions for all events in the file.
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    filename = file.filename.lower()

    try:
        if filename.endswith(".csv"):
            rows = _parse_csv(file)
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = _parse_excel(file)
        else:
            return jsonify({"error": "Only CSV or Excel (.xlsx) files are supported"}), 400
    except Exception as e:
        return jsonify({"error": f"Could not parse file: {str(e)}"}), 400

    results = []
    total_savings = 0
    total_co2 = 0
    total_food_waste = 0

    for i, row in enumerate(rows):
        gate = transparency_gate(row)
        if gate["status"] == "BLOCKED":
            results.append({
                "row": i + 1,
                "event_name": row.get("event_name", f"Row {i+1}"),
                "status": "BLOCKED",
                "reason": gate["reason"]
            })
            continue

        prediction = predict_attendance(row)
        economics = calculate_waste_cost(
            planned=int(row.get("planned_quantity", 0)),
            predicted=prediction["predicted_attendance"],
            event_type=row.get("event_type", "general"),
            cost_per_person=float(row.get("cost_per_person", 15))
        )

        total_savings += economics["total_savings_usd"]
        total_co2 += economics["co2_saved_kg"]
        total_food_waste += economics["food_waste_lbs"]

        results.append({
            "row": i + 1,
            "event_name": row.get("event_name", f"Event {i+1}"),
            "event_type": row.get("event_type"),
            "expected_signups": prediction["expected_signups"],
            "predicted_attendance": prediction["predicted_attendance"],
            "show_rate_pct": prediction["show_rate_pct"],
            "confidence_level": prediction["confidence_level"],
            "over_prepared_by": economics["over_prepared_by"],
            "total_savings_usd": economics["total_savings_usd"],
            "food_waste_lbs": economics["food_waste_lbs"],
            "co2_saved_kg": economics["co2_saved_kg"],
            "status": "OK"
        })

    return jsonify({
        "events": results,
        "summary": {
            "total_events": len(rows),
            "processed": len([r for r in results if r["status"] == "OK"]),
            "blocked": len([r for r in results if r["status"] == "BLOCKED"]),
            "total_savings_usd": round(total_savings, 2),
            "total_co2_saved_kg": round(total_co2, 2),
            "total_food_waste_lbs": round(total_food_waste, 2)
        }
    })


@app.route("/api/sample-csv")
def download_sample():
    """Serve the sample CSV file for download."""
    return send_file(
        os.path.join(os.path.dirname(__file__), "data", "sample_events.csv"),
        mimetype="text/csv",
        as_attachment=True,
        download_name="sample_events.csv"
    )


@app.route("/api/dataset/<name>")
def download_dataset(name):
    """Serve named dataset files."""
    allowed = {
        "sample": "sample_events.csv",
        "asu-spring": "asu_spring_semester.csv",
        "worst-case": "worst_case_events.csv",
        "best-case": "best_case_events.csv"
    }
    if name not in allowed:
        return jsonify({"error": "Dataset not found"}), 404
    filename = allowed[name]
    return send_file(
        os.path.join(os.path.dirname(__file__), "data", filename),
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename
    )


@app.route("/datasets")
def datasets():
    return render_template("datasets.html")


@app.route("/pitch")
def pitch():
    return send_file(
        os.path.join(os.path.dirname(__file__), "presentation", "index.html")
    )


@app.route("/presentation/<path:filename>")
def presentation_static(filename):
    return send_file(
        os.path.join(os.path.dirname(__file__), "presentation", filename)
    )


def _parse_csv(file) -> list:
    content = file.read().decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for row in reader:
        rows.append(_clean_row(row))
    return rows


def _parse_excel(file) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers = [str(cell.value).strip() for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            row_dict = dict(zip(headers, row))
            rows.append(_clean_row(row_dict))
    return rows


def _clean_row(row: dict) -> dict:
    """Normalize a CSV/Excel row to match the API payload format."""
    cleaned = {}
    for k, v in row.items():
        if v is None or str(v).strip() == "":
            continue
        cleaned[k.strip()] = str(v).strip()

    # Convert numeric fields
    for field in ["expected_signups", "planned_quantity"]:
        if field in cleaned:
            try:
                cleaned[field] = int(float(cleaned[field]))
            except ValueError:
                pass

    for field in ["cost_per_person", "interest_match_score", "historical_show_rate"]:
        if field in cleaned:
            try:
                cleaned[field] = float(cleaned[field])
            except ValueError:
                pass

    # Default planned_quantity to expected_signups if not provided
    if "planned_quantity" not in cleaned and "expected_signups" in cleaned:
        cleaned["planned_quantity"] = cleaned["expected_signups"]

    return cleaned

@app.route("/api/student-insights", methods=["POST"])
def student_insights():
    data = request.get_json()
    insights = get_student_insights(data)
    return jsonify(insights)


@app.route("/api/send-email", methods=["POST"])
def send_email():
    """
    Send a personalized reminder email to a student.
    Body: { to_email, subject, email_body }
    """
    data = request.get_json()
    to_email = data.get("to_email", "").strip()
    subject = data.get("subject", "EventWise Reminder")
    email_body = data.get("email_body", "")

    if not email_body:
        return jsonify({"status": "ERROR", "reason": "No email content provided"}), 400

    result = send_reminder_email(to_email, subject, email_body)

    if result["status"] == "ERROR":
        return jsonify(result), 400

    return jsonify(result)

if __name__ == "__main__":
    app.run(debug=True)
