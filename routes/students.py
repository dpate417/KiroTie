"""
Flask routes for student scoring and email features.

Endpoints:
  POST /api/students/score-upload   — Upload CSV/Excel, score all students for an event
  POST /api/students/send-emails    — Send emails to selected scored students
  GET  /api/students/sample-csv     — Download sample student CSV
"""
from __future__ import annotations

import csv
import io
import os

from flask import Blueprint, current_app, jsonify, request, send_file

from data.events_seed import get_event_by_id
from models.student import StudentRecord
from routes.middleware import require_auth
from services.student_scoring_engine import score_all_students
from services.email_service import send_to_selected

students_bp = Blueprint("students", __name__)

SAMPLE_CSV_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "sample_students.csv")


def _ok(data):
    return jsonify({"status": "ok", "data": data, "error": None}), 200


def _error(msg, code=400):
    return jsonify({"status": "error", "data": None, "error": msg}), code


# ── POST /api/students/score-upload ─────────────────────────────────────────

@students_bp.route("/students/score-upload", methods=["POST"])
@require_auth
def score_upload():
    """
    Upload a student CSV/Excel file and score all students for a given event.
    Form fields: file (required), event_id (required)
    """
    event_id = request.form.get("event_id", "").strip()
    if not event_id:
        return _error("event_id is required")

    event = get_event_by_id(event_id)
    if not event:
        return _error(f"Event '{event_id}' not found", 404)

    if "file" not in request.files:
        return _error("No file uploaded")

    file = request.files["file"]
    filename = file.filename.lower()

    try:
        if filename.endswith(".csv"):
            students = _parse_csv(file)
        elif filename.endswith((".xlsx", ".xls")):
            students = _parse_excel(file)
        else:
            return _error("Only CSV or Excel (.xlsx) files are supported")
    except Exception as e:
        return _error(f"Could not parse file: {e}")

    if not students:
        return _error("No student records found in file")

    scores = score_all_students(students, event)

    return _ok({
        "event_id": event_id,
        "event_name": event.name,
        "total": len(scores),
        "students": [_score_to_dict(s) for s in scores],
        "summary": {
            "high": sum(1 for s in scores if s.likelihood == "High"),
            "medium": sum(1 for s in scores if s.likelihood == "Medium"),
            "low": sum(1 for s in scores if s.likelihood == "Low"),
            "avg_score": round(sum(s.score for s in scores) / len(scores), 1) if scores else 0,
        }
    })


# ── POST /api/students/send-emails ──────────────────────────────────────────

@students_bp.route("/students/send-emails", methods=["POST"])
@require_auth
def send_emails():
    """
    Send personalized emails to selected students.
    Body: { students: [ {email, name, email_subject, email_body, has_valid_email}, ... ] }
    """
    body = request.get_json(silent=True) or {}
    selected = body.get("students", [])

    if not selected:
        return _error("No students provided")

    # Rebuild StudentScore objects from the payload
    from models.student import StudentScore
    score_objs = []
    for s in selected:
        score_objs.append(StudentScore(
            email=s.get("email", ""),
            name=s.get("name", ""),
            event_id=s.get("event_id", ""),
            score=s.get("score", 0),
            grade=s.get("grade", ""),
            likelihood=s.get("likelihood", ""),
            factors=s.get("factors", []),
            email_subject=s.get("email_subject", "EventWise Reminder"),
            email_body=s.get("email_body", ""),
            has_valid_email=s.get("has_valid_email", True),
        ))

    results = send_to_selected(score_objs)

    sent = sum(1 for r in results if r.status == "SENT")
    failed = sum(1 for r in results if r.status == "FAILED")
    skipped = sum(1 for r in results if r.status == "SKIPPED")

    return _ok({
        "summary": {"total": len(results), "sent": sent, "failed": failed, "skipped": skipped},
        "results": [{"email": r.email, "name": r.name, "status": r.status, "reason": r.reason} for r in results]
    })


# ── GET /api/students/sample-csv ────────────────────────────────────────────

@students_bp.route("/students/sample-csv", methods=["GET"])
def sample_csv():
    """Download the sample student CSV."""
    path = os.path.abspath(SAMPLE_CSV_PATH)
    if not os.path.exists(path):
        return _error("Sample CSV not found", 404)
    return send_file(path, mimetype="text/csv", as_attachment=True, download_name="sample_students.csv")


# ── GET /api/students/events ─────────────────────────────────────────────────

@students_bp.route("/students/events", methods=["GET"])
@require_auth
def list_events_for_scoring():
    """Return list of events available for student scoring."""
    from data.events_seed import EVENTS
    return _ok([{"id": e.id, "name": e.name, "date": e.date, "time": e.time,
                 "location": e.location, "signup_count": e.signup_count} for e in EVENTS])


# ── Helpers ──────────────────────────────────────────────────────────────────

def _score_to_dict(s) -> dict:
    return {
        "email": s.email,
        "name": s.name,
        "event_id": s.event_id,
        "score": s.score,
        "grade": s.grade,
        "likelihood": s.likelihood,
        "factors": s.factors,
        "email_subject": s.email_subject,
        "email_body": s.email_body,
        "has_valid_email": s.has_valid_email,
    }


def _parse_csv(file) -> list[StudentRecord]:
    content = file.read().decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    return [_row_to_student(row) for row in reader if any(v.strip() for v in row.values())]


def _parse_excel(file) -> list[StudentRecord]:
    import openpyxl
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers = [str(c.value).strip() for c in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            records.append(_row_to_student(dict(zip(headers, row))))
    return records


def _row_to_student(row: dict) -> StudentRecord:
    def _int(k, default=0):
        try:
            return int(float(str(row.get(k, default) or default)))
        except (ValueError, TypeError):
            return default

    def _float(k, default=0.0):
        try:
            return float(str(row.get(k, default) or default))
        except (ValueError, TypeError):
            return default

    def _str(k):
        return str(row.get(k, "") or "").strip()

    tags_raw = _str("interest_tags")
    tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

    attended = _int("events_attended")
    registered = _int("events_registered") or (attended + _int("events_missed"))

    return StudentRecord(
        email=_str("email"),
        name=_str("student_name") or _str("name"),
        major=_str("major") or _str("student_major"),
        events_attended=attended,
        events_registered=max(registered, attended),
        events_missed=_int("events_missed"),
        avg_registration_days_before=_float("avg_registration_days_before", 3.0),
        interest_tags=tags,
    )
